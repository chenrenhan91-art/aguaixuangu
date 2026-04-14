import csv
import io
import json
import re
import sqlite3
import uuid
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

from fastapi import HTTPException, UploadFile

from app.core.config import get_settings
from app.schemas.trade_diagnostics import (
    DiagnosticInsight,
    ImportBatchSummary,
    ImportProfile,
    ImportProfilesResponse,
    MetricCard,
    StandardFieldDefinition,
    TradeDiagnosticsAIResponse,
    TradeDiagnosticsResponse,
    TradeImportResponse,
    TradeStyleProfile,
)
from app.services.gemini_client import GeminiClientError, call_gemini_trade_diagnostics_analysis


SUPPORTED_IMPORT_PROFILES = [
    ImportProfile(
        profile_id="eastmoney_excel",
        display_name="东方财富证券导出",
        broker="东方财富证券",
        supported_extensions=[".xlsx", ".xls", ".csv"],
        recommended_format="Excel / CSV",
        description="优先支持券商客户端导出的交割单或历史成交表格，近 1 个月以上记录即可生成更有效诊断。",
        export_steps=[
            "在东方财富证券交易端打开历史成交或交割单查询。",
            "选择完整时间范围后导出 Excel 或 CSV。",
            "建议至少导出近 1 个月记录。",
            "直接上传导出的文件，无需截图。",
        ],
    ),
    ImportProfile(
        profile_id="ths_generic",
        display_name="同花顺/券商交易端导出",
        broker="同花顺系交易端",
        supported_extensions=[".xlsx", ".xls", ".csv"],
        recommended_format="Excel / CSV",
        description="适配常见的同花顺交易端与券商联营客户端导出文件，近 1 个月以上记录即可生成更有效诊断。",
        export_steps=[
            "在交易端进入历史成交、资金流水或交割单页面。",
            "导出为 Excel 或 CSV，保留原始列名。",
            "建议至少导出近 1 个月记录。",
            "上传文件后系统会自动识别代码、买卖方向和费用字段。",
        ],
    ),
    ImportProfile(
        profile_id="huatai",
        display_name="华泰证券导出",
        broker="华泰证券",
        supported_extensions=[".xlsx", ".xls", ".csv"],
        recommended_format="Excel / CSV",
        description="适配以交割流水和历史成交为主的常见表头，近 1 个月以上记录即可生成更有效诊断。",
        export_steps=[
            "从华泰证券客户端导出历史成交或交割流水。",
            "建议选择包含费用和净发生金额的明细表。",
            "建议至少导出近 1 个月记录。",
            "上传后优先按最新批次生成交易诊断。",
        ],
    ),
    ImportProfile(
        profile_id="generic_csv",
        display_name="通用 CSV / Excel",
        broker="其他券商",
        supported_extensions=[".xlsx", ".xls", ".csv"],
        recommended_format="CSV 优先",
        description="适合其他券商或已自行整理过列名的交易流水文件，近 1 个月以上记录即可生成更有效诊断。",
        export_steps=[
            "保留每笔成交记录，不要只保留汇总行。",
            "建议至少覆盖近 1 个月交易记录。",
            "确保至少包含日期、代码、方向、数量、价格。",
            "费用字段若缺失也可先导入，系统会按零值处理。",
        ],
    ),
]

STANDARD_FIELDS = [
    StandardFieldDefinition(field_name="trade_date", display_name="成交日期", required=True, description="支持 YYYY-MM-DD、YYYY/MM/DD、YYYYMMDD。"),
    StandardFieldDefinition(field_name="trade_time", display_name="成交时间", required=False, description="可为空，若无则按批次行号排序。"),
    StandardFieldDefinition(field_name="symbol", display_name="证券代码", required=True, description="系统会自动补齐到 6 位代码。"),
    StandardFieldDefinition(field_name="stock_name", display_name="证券名称", required=False, description="用于展示与复盘。"),
    StandardFieldDefinition(field_name="side", display_name="买卖方向", required=True, description="支持 买入 / 卖出 / buy / sell 等常见写法。"),
    StandardFieldDefinition(field_name="quantity", display_name="成交数量", required=True, description="整数股数。"),
    StandardFieldDefinition(field_name="price", display_name="成交价格", required=True, description="单笔成交均价或成交价。"),
    StandardFieldDefinition(field_name="amount", display_name="成交金额", required=False, description="缺失时会自动按价格与数量补算。"),
    StandardFieldDefinition(field_name="commission", display_name="佣金", required=False, description="默认按 0 处理。"),
    StandardFieldDefinition(field_name="stamp_tax", display_name="印花税", required=False, description="默认按 0 处理。"),
    StandardFieldDefinition(field_name="transfer_fee", display_name="过户费", required=False, description="默认按 0 处理。"),
    StandardFieldDefinition(field_name="other_fee", display_name="其他费用", required=False, description="包含规费等其他费用。"),
    StandardFieldDefinition(field_name="net_amount", display_name="净发生金额", required=False, description="若缺失会自动推断。"),
]

HEADER_ALIASES = {
    "trade_date": ["成交日期", "交易日期", "成交时间", "日期", "发生日期", "委托日期", "业务日期"],
    "trade_time": ["成交时间", "成交时刻", "时间"],
    "symbol": ["证券代码", "股票代码", "代码", "证券代号"],
    "stock_name": ["证券名称", "股票名称", "名称", "证券简称"],
    "side": ["买卖方向", "买卖标志", "业务名称", "方向", "操作"],
    "quantity": ["成交数量", "数量", "成交股数", "发生数量", "成交份额"],
    "price": ["成交价格", "成交均价", "均价", "价格"],
    "amount": ["成交金额", "发生金额", "清算金额", "金额", "发生发生金额"],
    "commission": ["佣金", "净佣金", "手续费", "交易佣金"],
    "stamp_tax": ["印花税"],
    "transfer_fee": ["过户费"],
    "other_fee": ["其他费用", "风险金", "规费", "杂费"],
    "net_amount": ["净发生金额", "清算金额", "净收付", "清算净额", "发生净额"],
    "account_masked": ["股东代码", "股东账户", "资金账户", "账户"],
    "business_flag": ["业务标志", "业务类型"],
    "security_type": ["证券类别"],
}

BUY_TOKENS = ("买", "买入", "证券买入", "担保品买入", "buy")
SELL_TOKENS = ("卖", "卖出", "证券卖出", "担保品卖出", "sell")


@dataclass
class StandardizedTrade:
    trade_date: date
    trade_time: Optional[str]
    symbol: str
    stock_name: str
    market: str
    side: str
    quantity: int
    price: float
    amount: float
    commission: float
    stamp_tax: float
    transfer_fee: float
    other_fee: float
    net_amount: float
    account_masked: str
    broker: str
    source_type: str
    raw_row_id: int


def _settings():
    return get_settings()


def _db_path() -> Path:
    path = _settings().data_dir / "processed" / "trade_diagnostics.sqlite3"
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _imports_dir() -> Path:
    path = _settings().data_dir / "raw" / "trade_imports"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _trade_ai_cache_path() -> Path:
    path = _settings().data_dir / "processed" / "ai_trade_diagnostics_cache.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _load_trade_ai_cache() -> dict[str, Any]:
    path = _trade_ai_cache_path()
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def _get_cached_trade_ai_analysis(batch_id: str) -> Optional[dict[str, Any]]:
    return _load_trade_ai_cache().get(batch_id)


def _save_cached_trade_ai_analysis(batch_id: str, payload: dict[str, Any]) -> None:
    cache = _load_trade_ai_cache()
    cache[batch_id] = payload
    _trade_ai_cache_path().write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


@contextmanager
def _connect() -> Iterable[sqlite3.Connection]:
    conn = sqlite3.connect(_db_path())
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def ensure_trade_tables() -> None:
    with _connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS import_batches (
                batch_id TEXT PRIMARY KEY,
                imported_at TEXT NOT NULL,
                broker TEXT NOT NULL,
                source_type TEXT NOT NULL,
                filename TEXT NOT NULL,
                detected_format TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                imported_count INTEGER NOT NULL,
                ignored_count INTEGER NOT NULL,
                symbol_count INTEGER NOT NULL,
                start_date TEXT,
                end_date TEXT,
                notes TEXT
            );

            CREATE TABLE IF NOT EXISTS trades (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id TEXT NOT NULL,
                trade_date TEXT NOT NULL,
                trade_time TEXT,
                symbol TEXT NOT NULL,
                stock_name TEXT,
                market TEXT,
                side TEXT NOT NULL,
                quantity INTEGER NOT NULL,
                price REAL NOT NULL,
                amount REAL NOT NULL,
                commission REAL NOT NULL DEFAULT 0,
                stamp_tax REAL NOT NULL DEFAULT 0,
                transfer_fee REAL NOT NULL DEFAULT 0,
                other_fee REAL NOT NULL DEFAULT 0,
                net_amount REAL NOT NULL,
                account_masked TEXT,
                broker TEXT NOT NULL,
                source_type TEXT NOT NULL,
                raw_row_id INTEGER,
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_trades_batch_date ON trades(batch_id, trade_date);
            CREATE INDEX IF NOT EXISTS idx_trades_symbol_date ON trades(symbol, trade_date);
            """
        )


def get_import_profiles() -> ImportProfilesResponse:
    return ImportProfilesResponse(profiles=SUPPORTED_IMPORT_PROFILES, standard_fields=STANDARD_FIELDS)


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\s_]+", "", str(value or "")).strip().lower()


def _build_header_mapping(headers: list[str]) -> dict[str, str]:
    normalized = {_normalize_header(header): header for header in headers}
    mapping = {}
    for standard_field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            candidate = normalized.get(_normalize_header(alias))
            if candidate:
                mapping[standard_field] = candidate
                break
    return mapping


def _parse_float(value: Any, default: float = 0.0) -> float:
    if value in (None, "", "--"):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(",", "").replace("￥", "").replace("¥", "")
    cleaned = cleaned.replace("元", "").replace("股", "")
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    try:
        return float(cleaned)
    except ValueError:
        return default


def _parse_int(value: Any) -> int:
    return int(round(_parse_float(value, default=0.0)))


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, int):
        if 19000101 <= value <= 29991231:
            return datetime.strptime(str(value), "%Y%m%d").date()
    raw = str(value or "").strip()
    if not raw:
        raise ValueError("missing trade date")
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"unsupported date format: {raw}")


def _parse_time(value: Any) -> Optional[str]:
    if value in (None, "", "--"):
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, int):
        raw_digits = str(value).zfill(6)[-6:]
        return f"{raw_digits[0:2]}:{raw_digits[2:4]}:{raw_digits[4:6]}"
    raw = str(value).strip()
    if len(raw) >= 8 and ":" in raw:
        return raw[:8]
    if len(raw) == 6 and raw.isdigit():
        return f"{raw[0:2]}:{raw[2:4]}:{raw[4:6]}"
    return None


def _parse_symbol(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if not digits:
        raise ValueError("missing symbol")
    return digits[-6:].zfill(6)


def _infer_market(symbol: str) -> str:
    if symbol.startswith(("60", "68", "90")):
        return "SH"
    if symbol.startswith(("00", "30", "20")):
        return "SZ"
    if symbol.startswith("8"):
        return "BJ"
    return ""


def _parse_side(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if any(token in raw for token in BUY_TOKENS):
        return "buy"
    if any(token in raw for token in SELL_TOKENS):
        return "sell"
    raise ValueError(f"unsupported side: {value}")


def _read_csv_rows(content: bytes) -> list[dict[str, Any]]:
    encodings = ("utf-8-sig", "utf-8", "gb18030", "gbk")
    last_error = None
    for encoding in encodings:
        try:
            text = content.decode(encoding)
            sample = text[:2048]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
            except csv.Error:
                dialect = csv.excel
            reader = csv.DictReader(io.StringIO(text), dialect=dialect)
            return [dict(row) for row in reader]
        except UnicodeDecodeError as exc:
            last_error = exc
    raise HTTPException(status_code=400, detail=f"无法识别 CSV 编码: {last_error}")


def _read_excel_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="当前环境未安装 openpyxl，暂时无法解析 Excel。") from exc

    workbook = load_workbook(filename=io.BytesIO(content), data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_index = 0
    for index, row in enumerate(rows):
        if row and sum(1 for value in row if value not in (None, "")) >= 3:
            header_index = index
            break

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[header_index]]
    data_rows = []
    for row in rows[header_index + 1 :]:
        if not row or all(cell in (None, "") for cell in row):
            continue
        data_rows.append(
            {
                headers[idx]: row[idx] if idx < len(row) else None
                for idx in range(len(headers))
                if headers[idx]
            }
        )
    return data_rows


def _read_legacy_excel_rows(content: bytes) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="当前环境未安装 xlrd，暂时无法解析 XLS。") from exc

    workbook = xlrd.open_workbook(file_contents=content)
    worksheet = workbook.sheet_by_index(0)
    if worksheet.nrows == 0:
        return []

    header_index = 0
    for index in range(worksheet.nrows):
        row = worksheet.row_values(index)
        if row and sum(1 for value in row if value not in (None, "")) >= 3:
            header_index = index
            break

    headers = [str(cell).strip() if cell is not None else "" for cell in worksheet.row_values(header_index)]
    data_rows: list[dict[str, Any]] = []
    for row_idx in range(header_index + 1, worksheet.nrows):
        row = worksheet.row_values(row_idx)
        if not row or all(cell in (None, "") for cell in row):
            continue
        data_rows.append(
            {
                headers[idx]: row[idx] if idx < len(row) else None
                for idx in range(len(headers))
                if headers[idx]
            }
        )
    return data_rows


def _load_tabular_rows(filename: str, content: bytes) -> tuple[str, list[dict[str, Any]]]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".csv", ".txt"}:
        return "csv", _read_csv_rows(content)
    if suffix in {".xlsx"}:
        return "excel", _read_excel_rows(content)
    if suffix in {".xls"}:
        return "excel", _read_legacy_excel_rows(content)
    raise HTTPException(status_code=400, detail="当前仅支持 CSV、XLSX 与 XLS 文件导入。")


def _standardize_rows(
    rows: list[dict[str, Any]], broker: str, source_type: str
) -> tuple[list[StandardizedTrade], list[str]]:
    if not rows:
        raise HTTPException(status_code=400, detail="导入文件没有可识别的数据行。")

    headers = [key for key in rows[0].keys() if key]
    mapping = _build_header_mapping(headers)
    required_fields = ("trade_date", "symbol", "side", "quantity", "price")
    missing = [field for field in required_fields if field not in mapping]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"文件缺少关键字段映射: {', '.join(missing)}",
        )

    standardized = []
    ignored_rows = []
    for index, row in enumerate(rows, start=1):
        try:
            business_flag = str(row.get(mapping.get("business_flag"), "") or "").strip()
            security_type = str(row.get(mapping.get("security_type"), "") or "").strip()
            if business_flag and not any(token in business_flag.lower() for token in BUY_TOKENS + SELL_TOKENS):
                ignored_rows.append(f"row-{index}")
                continue
            if "指定交易" in security_type:
                ignored_rows.append(f"row-{index}")
                continue
            trade_date = _parse_date(row.get(mapping["trade_date"]))
            symbol = _parse_symbol(row.get(mapping["symbol"]))
            quantity = abs(_parse_int(row.get(mapping["quantity"])))
            price = _parse_float(row.get(mapping["price"]))
            side = _parse_side(row.get(mapping["side"]))
            stock_name = str(row.get(mapping.get("stock_name"), "") or "").strip()
            trade_time = _parse_time(row.get(mapping.get("trade_time")))
            commission = _parse_float(row.get(mapping.get("commission")))
            stamp_tax = _parse_float(row.get(mapping.get("stamp_tax")))
            transfer_fee = _parse_float(row.get(mapping.get("transfer_fee")))
            other_fee = _parse_float(row.get(mapping.get("other_fee")))
            amount = _parse_float(row.get(mapping.get("amount")))
            if amount <= 0:
                amount = round(quantity * price, 2)
            net_amount = _parse_float(row.get(mapping.get("net_amount")))
            if net_amount == 0:
                fee_total = commission + stamp_tax + transfer_fee + other_fee
                signed_amount = amount if side == "sell" else -amount
                signed_fee = -fee_total
                net_amount = round(signed_amount + signed_fee, 2)
            if quantity <= 0 or price <= 0 or (amount <= 0 and abs(net_amount) < 1e-8):
                ignored_rows.append(f"row-{index}")
                continue
            standardized.append(
                StandardizedTrade(
                    trade_date=trade_date,
                    trade_time=trade_time,
                    symbol=symbol,
                    stock_name=stock_name,
                    market=_infer_market(symbol),
                    side=side,
                    quantity=quantity,
                    price=price,
                    amount=amount,
                    commission=commission,
                    stamp_tax=stamp_tax,
                    transfer_fee=transfer_fee,
                    other_fee=other_fee,
                    net_amount=net_amount,
                    account_masked=str(row.get(mapping.get("account_masked"), "") or "").strip(),
                    broker=broker,
                    source_type=source_type,
                    raw_row_id=index,
                )
            )
        except ValueError:
            ignored_rows.append(f"row-{index}")

    if not standardized:
        raise HTTPException(status_code=400, detail="没有可导入的有效交易行，请检查文件内容。")

    standardized.sort(key=lambda item: (item.trade_date, item.trade_time or "", item.raw_row_id))
    return standardized, ignored_rows


def _row_to_batch_summary(row: sqlite3.Row) -> ImportBatchSummary:
    return ImportBatchSummary(
        batch_id=row["batch_id"],
        imported_at=datetime.fromisoformat(row["imported_at"]),
        broker=row["broker"],
        source_type=row["source_type"],
        filename=row["filename"],
        detected_format=row["detected_format"],
        row_count=row["row_count"],
        imported_count=row["imported_count"],
        ignored_count=row["ignored_count"],
        symbol_count=row["symbol_count"],
        start_date=date.fromisoformat(row["start_date"]) if row["start_date"] else None,
        end_date=date.fromisoformat(row["end_date"]) if row["end_date"] else None,
        notes=row["notes"],
    )


async def import_trade_file(upload: UploadFile, profile_id: str) -> TradeImportResponse:
    ensure_trade_tables()
    profile = next((item for item in SUPPORTED_IMPORT_PROFILES if item.profile_id == profile_id), None)
    if profile is None:
        raise HTTPException(status_code=404, detail="未知导入模板。")
    if not upload.filename:
        raise HTTPException(status_code=400, detail="请先选择导入文件。")

    content = await upload.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空。")

    detected_format, rows = _load_tabular_rows(upload.filename, content)
    source_type = "excel" if detected_format == "excel" else "csv"
    trades, ignored_rows = _standardize_rows(rows, broker=profile.broker, source_type=source_type)

    batch_id = uuid.uuid4().hex[:12]
    imported_at = datetime.now().isoformat()
    target_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{batch_id}{Path(upload.filename).suffix.lower()}"
    (_imports_dir() / target_name).write_bytes(content)

    start_date = min(item.trade_date for item in trades).isoformat()
    end_date = max(item.trade_date for item in trades).isoformat()
    notes = f"模板 {profile.display_name} 自动映射导入，共识别 {len(trades)} 条有效记录。"

    with _connect() as conn:
        conn.execute(
            """
            INSERT INTO import_batches (
                batch_id, imported_at, broker, source_type, filename, detected_format,
                row_count, imported_count, ignored_count, symbol_count, start_date, end_date, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                batch_id,
                imported_at,
                profile.broker,
                source_type,
                upload.filename,
                detected_format,
                len(rows),
                len(trades),
                len(ignored_rows),
                len({item.symbol for item in trades}),
                start_date,
                end_date,
                notes,
            ),
        )
        conn.executemany(
            """
            INSERT INTO trades (
                batch_id, trade_date, trade_time, symbol, stock_name, market, side,
                quantity, price, amount, commission, stamp_tax, transfer_fee, other_fee,
                net_amount, account_masked, broker, source_type, raw_row_id, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    batch_id,
                    item.trade_date.isoformat(),
                    item.trade_time,
                    item.symbol,
                    item.stock_name,
                    item.market,
                    item.side,
                    item.quantity,
                    item.price,
                    item.amount,
                    item.commission,
                    item.stamp_tax,
                    item.transfer_fee,
                    item.other_fee,
                    item.net_amount,
                    item.account_masked,
                    item.broker,
                    item.source_type,
                    item.raw_row_id,
                    imported_at,
                )
                for item in trades
            ],
        )

    batch = ImportBatchSummary(
        batch_id=batch_id,
        imported_at=datetime.fromisoformat(imported_at),
        broker=profile.broker,
        source_type=source_type,
        filename=upload.filename,
        detected_format=detected_format,
        row_count=len(rows),
        imported_count=len(trades),
        ignored_count=len(ignored_rows),
        symbol_count=len({item.symbol for item in trades}),
        start_date=date.fromisoformat(start_date),
        end_date=date.fromisoformat(end_date),
        notes=notes,
    )
    return TradeImportResponse(batch=batch, message="导入完成，已生成最新交易诊断。")


def _load_recent_batches(limit: int = 5) -> list[ImportBatchSummary]:
    ensure_trade_tables()
    with _connect() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM import_batches
            ORDER BY imported_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [_row_to_batch_summary(row) for row in rows]


def _load_latest_batch() -> Optional[ImportBatchSummary]:
    batches = _load_recent_batches(limit=1)
    return batches[0] if batches else None


def _load_batch_by_id(batch_id: str) -> Optional[ImportBatchSummary]:
    ensure_trade_tables()
    with _connect() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM import_batches
            WHERE batch_id = ?
            """,
            (batch_id,),
        ).fetchone()
    return _row_to_batch_summary(row) if row is not None else None


def _load_trades_for_batch(batch_id: str) -> list[dict[str, Any]]:
    with _connect() as conn:
        rows = conn.execute(
            """
            SELECT *
            FROM trades
            WHERE batch_id = ?
            ORDER BY trade_date ASC, COALESCE(trade_time, ''), raw_row_id ASC, id ASC
            """,
            (batch_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def _build_demo_diagnostics() -> TradeDiagnosticsResponse:
    return TradeDiagnosticsResponse(
        status="demo",
        account_label="演示账户",
        coverage_text="尚未导入真实交割单，当前为示例分析视图。",
        latest_batch=None,
        summary_metrics=[
            MetricCard(label="闭环交易数", value="36", detail="示例数据中的已完成买卖配对。"),
            MetricCard(label="胜率", value="52.8%", detail="盈利单占比略高，但稳定性一般。"),
            MetricCard(label="盈亏比", value="1.34", detail="单笔盈利略大于单笔亏损。"),
            MetricCard(label="平均持仓", value="4.6 天", detail="更接近 3-8 天的短波段打法。"),
        ],
        style_profile=TradeStyleProfile(
            style_id="swing_short",
            display_name="短波段交易型",
            confidence=0.76,
            summary="盈利主要来自趋势延续阶段的 3 到 8 天持有，频繁换股会明显拉低表现。",
            traits=["偏好热点行业龙头", "更适合确认后跟随", "不适合高频冲动交易"],
        ),
        win_loss_comparison=[
            MetricCard(label="盈利单平均持有", value="6.1 天", detail="赚钱时更愿意耐心持有。"),
            MetricCard(label="亏损单平均持有", value="2.4 天", detail="止损尚算及时，但容易追高后回撤。"),
            MetricCard(label="盈利单平均收益", value="+8.2%", detail="大多来自顺势延续与二次突破。"),
            MetricCard(label="亏损单平均回撤", value="-5.9%", detail="集中在情绪高位接力与弱势环境出手。"),
        ],
        error_patterns=[
            DiagnosticInsight(title="热点追高后回落", detail="亏损单多数发生在放量冲高次日才介入，买点偏后。", severity="medium"),
            DiagnosticInsight(title="赚小亏大偶发出现", detail="部分盈利单在第一波拉升后过早止盈，反而保留了弱势单。", severity="medium"),
        ],
        effective_patterns=[
            DiagnosticInsight(title="板块共振时顺势跟随有效", detail="当行业主线明确、成交额同步放大时，交易胜率更高。", severity="positive"),
            DiagnosticInsight(title="首笔决策质量较高", detail="单次买入就获利的案例明显优于反复补仓的案例。", severity="positive"),
        ],
        recommendations=[
            "把主要精力放在 3 到 8 天的趋势延续交易，不必追求日内频繁切换。",
            "热点股只做第一次回踩确认，不在连续加速后再追入。",
            "对已经跌破买入逻辑的位置设硬性离场，避免把短线单拖成长线。 ",
        ],
        recent_batches=[],
    )


def _compute_round_trips(trades: list[dict[str, Any]]) -> list[dict[str, Any]]:
    positions: dict[str, list[dict[str, Any]]] = {}
    round_trips: list[dict[str, Any]] = []

    for trade in trades:
        symbol = trade["symbol"]
        positions.setdefault(symbol, [])
        fees = (
            float(trade.get("commission") or 0)
            + float(trade.get("stamp_tax") or 0)
            + float(trade.get("transfer_fee") or 0)
            + float(trade.get("other_fee") or 0)
        )
        quantity = int(trade["quantity"])
        if quantity <= 0:
            continue

        if trade["side"] == "buy":
            positions[symbol].append(
                {
                    "remaining_qty": quantity,
                    "entry_date": date.fromisoformat(trade["trade_date"]),
                    "entry_price": float(trade["price"]),
                    "fee_per_share": fees / quantity,
                    "stock_name": trade.get("stock_name") or symbol,
                }
            )
            continue

        remaining_sell_qty = quantity
        sell_fee_per_share = fees / quantity
        exit_date = date.fromisoformat(trade["trade_date"])
        while remaining_sell_qty > 0 and positions[symbol]:
            buy_lot = positions[symbol][0]
            matched_qty = min(remaining_sell_qty, buy_lot["remaining_qty"])
            pnl = (
                (float(trade["price"]) - buy_lot["entry_price"]) * matched_qty
                - (buy_lot["fee_per_share"] + sell_fee_per_share) * matched_qty
            )
            entry_cost = buy_lot["entry_price"] * matched_qty + buy_lot["fee_per_share"] * matched_qty
            holding_days = max((exit_date - buy_lot["entry_date"]).days, 0)
            round_trips.append(
                {
                    "symbol": symbol,
                    "stock_name": buy_lot["stock_name"],
                    "entry_date": buy_lot["entry_date"],
                    "exit_date": exit_date,
                    "matched_qty": matched_qty,
                    "pnl": pnl,
                    "return_pct": pnl / entry_cost if entry_cost else 0.0,
                    "holding_days": holding_days,
                }
            )
            buy_lot["remaining_qty"] -= matched_qty
            remaining_sell_qty -= matched_qty
            if buy_lot["remaining_qty"] == 0:
                positions[symbol].pop(0)

    return round_trips


def _format_percent(value: float) -> str:
    return f"{value * 100:.1f}%"


def _infer_style(round_trips: list[dict[str, Any]], trades: list[dict[str, Any]]) -> TradeStyleProfile:
    closed_count = len(round_trips)
    if closed_count == 0:
        return TradeStyleProfile(
            style_id="unclassified",
            display_name="待分类",
            confidence=0.35,
            summary="当前数据暂时无法形成完整买卖闭环，先继续导入更完整的交割单。",
            traits=["建议至少覆盖 1 个月以上交易记录"],
        )

    average_holding = sum(item["holding_days"] for item in round_trips) / closed_count
    active_days = {
        date.fromisoformat(trade["trade_date"]) if isinstance(trade["trade_date"], str) else trade["trade_date"]
        for trade in trades
    }
    trades_per_day = len(trades) / max(len(active_days), 1)

    if average_holding <= 2.5 and trades_per_day >= 2.2:
        return TradeStyleProfile(
            style_id="high_frequency_short",
            display_name="高频短线型",
            confidence=0.79,
            summary="交易节奏较快，偏向短线试错与快速切换，收益更依赖出手节奏而非持有耐心。",
            traits=["出手频率高", "更吃执行纪律", "手续费侵蚀需要重点控制"],
        )
    if average_holding <= 8:
        return TradeStyleProfile(
            style_id="swing_short",
            display_name="短波段交易型",
            confidence=0.82,
            summary="交易结果更像 3 到 8 天的顺势波段，适合围绕主线做确认后的跟随。",
            traits=["偏顺势而为", "适合板块共振时参与", "不宜追求过多日内切换"],
        )
    return TradeStyleProfile(
        style_id="trend_hold",
        display_name="耐心波段型",
        confidence=0.74,
        summary="持仓更愿意跨越多个交易日等待趋势展开，风格偏中短趋势而非短线博弈。",
        traits=["更依赖趋势结构", "适合减少高频换股", "应强化止损与加仓规则"],
    )


def _build_metric_cards(round_trips: list[dict[str, Any]], trades: list[dict[str, Any]]) -> tuple[list[MetricCard], dict[str, float]]:
    closed_count = len(round_trips)
    if closed_count == 0:
        return (
            [
                MetricCard(label="闭环交易数", value="0", detail="当前还没有可配对的完整买卖记录。"),
                MetricCard(label="胜率", value="--", detail="需要至少一笔闭环交易。"),
                MetricCard(label="盈亏比", value="--", detail="需要更多数据。"),
                MetricCard(label="平均持仓", value="--", detail="需要更多数据。"),
            ],
            {},
        )

    wins = [item for item in round_trips if item["pnl"] > 0]
    losses = [item for item in round_trips if item["pnl"] <= 0]
    gross_profit = sum(item["pnl"] for item in wins)
    gross_loss = abs(sum(item["pnl"] for item in losses))
    avg_holding = sum(item["holding_days"] for item in round_trips) / closed_count
    win_rate = len(wins) / closed_count
    avg_win_pct = sum(item["return_pct"] for item in wins) / max(len(wins), 1)
    avg_loss_pct = sum(item["return_pct"] for item in losses) / max(len(losses), 1) if losses else 0.0
    profit_factor = gross_profit / gross_loss if gross_loss else gross_profit

    metrics = {
        "win_rate": win_rate,
        "avg_holding": avg_holding,
        "avg_win_pct": avg_win_pct,
        "avg_loss_pct": avg_loss_pct,
        "profit_factor": profit_factor,
        "closed_count": float(closed_count),
        "winning_hold": sum(item["holding_days"] for item in wins) / max(len(wins), 1),
        "losing_hold": sum(item["holding_days"] for item in losses) / max(len(losses), 1),
    }

    cards = [
        MetricCard(label="闭环交易数", value=str(closed_count), detail="已按 FIFO 方式完成买卖配对。"),
        MetricCard(label="胜率", value=_format_percent(win_rate), detail="统计盈利交易在全部闭环交易中的占比。"),
        MetricCard(label="盈亏比", value=f"{profit_factor:.2f}", detail="总盈利 / 总亏损，越高越说明收益质量更好。"),
        MetricCard(label="平均持仓", value=f"{avg_holding:.1f} 天", detail="帮助判断你更适合短线、波段还是趋势持有。"),
    ]
    return cards, metrics


def _build_win_loss_comparison(round_trips: list[dict[str, Any]], metrics: dict[str, float]) -> list[MetricCard]:
    wins = [item for item in round_trips if item["pnl"] > 0]
    losses = [item for item in round_trips if item["pnl"] <= 0]
    avg_win = sum(item["return_pct"] for item in wins) / max(len(wins), 1)
    avg_loss = sum(item["return_pct"] for item in losses) / max(len(losses), 1) if losses else 0.0
    return [
        MetricCard(label="盈利单平均持有", value=f"{metrics.get('winning_hold', 0):.1f} 天", detail="赚钱时能否拿住，决定你的上沿收益。"),
        MetricCard(label="亏损单平均持有", value=f"{metrics.get('losing_hold', 0):.1f} 天", detail="亏损持有过久通常意味着离场纪律不足。"),
        MetricCard(label="盈利单平均收益", value=_format_percent(avg_win), detail="观察赚钱的典型回报区间。"),
        MetricCard(label="亏损单平均回撤", value=_format_percent(avg_loss), detail="帮助判断是否经常陷入赚小亏大。"),
    ]


def _build_insights(metrics: dict[str, float], trades: list[dict[str, Any]]) -> tuple[list[DiagnosticInsight], list[DiagnosticInsight], list[str]]:
    errors: list[DiagnosticInsight] = []
    strengths: list[DiagnosticInsight] = []
    recommendations: list[str] = []

    if not metrics:
        return (
            [DiagnosticInsight(title="闭环交易不足", detail="目前还无法形成可靠画像，建议导入更完整的历史流水。", severity="medium")],
            [],
            ["先覆盖至少 1 个月交易记录，再做风格与纪律分析。"],
        )

    buy_count = sum(1 for item in trades if item["side"] == "buy")
    sell_count = sum(1 for item in trades if item["side"] == "sell")
    multi_buy_ratio = buy_count / max(sell_count, 1)

    if metrics["avg_win_pct"] < abs(metrics["avg_loss_pct"]):
        errors.append(
            DiagnosticInsight(
                title="赚小亏大倾向",
                detail="平均盈利幅度仍小于平均亏损幅度，意味着止盈偏早或止损偏慢。",
                severity="high",
            )
        )
        recommendations.append("把第一止盈设置得更客观一些，避免盈利单过早离场。")

    if metrics["losing_hold"] > metrics["winning_hold"] * 1.25 and metrics["losing_hold"] >= 3:
        errors.append(
            DiagnosticInsight(
                title="亏损单持有偏久",
                detail="亏损交易比盈利交易更久，说明容易在逻辑失效后继续拖单。",
                severity="high",
            )
        )
        recommendations.append("把跌破关键均线或事件失效作为硬性离场条件。")

    if multi_buy_ratio > 1.35:
        errors.append(
            DiagnosticInsight(
                title="补仓摊薄使用偏多",
                detail="买入次数相对卖出次数偏高，说明更常通过加仓摊低成本来处理亏损单。",
                severity="medium",
            )
        )
        recommendations.append("只有在逻辑被二次确认时才允许加仓，避免用补仓替代认错。")

    if metrics["profit_factor"] >= 1.3:
        strengths.append(
            DiagnosticInsight(
                title="收益质量仍有基础",
                detail="总盈利与总亏损比值保持在可优化区间，说明策略并非失效，只是纪律还可继续打磨。",
                severity="positive",
            )
        )

    if metrics["winning_hold"] >= 3 and metrics["winning_hold"] > metrics["losing_hold"]:
        strengths.append(
            DiagnosticInsight(
                title="盈利单拿得住",
                detail="赚钱的交易能获得更长持有周期，这说明顺势持有对你有效。",
                severity="positive",
            )
        )
        recommendations.append("优先保留 3 到 8 天的趋势延续型交易，减少低质量频繁试错。")

    if metrics["win_rate"] < 0.45:
        recommendations.append("把交易触发条件再收紧一档，只做主线行业与事件共振更明确的机会。")
    else:
        recommendations.append("继续围绕高胜率场景建立清单，避免偏离最有效的交易模板。")

    return errors[:3], strengths[:3], recommendations[:4]


def _round_trip_symbol_summary(round_trips: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    by_symbol: dict[str, dict[str, Any]] = {}
    for item in round_trips:
        symbol = item["symbol"]
        symbol_payload = by_symbol.setdefault(
            symbol,
            {
                "symbol": symbol,
                "stock_name": item.get("stock_name") or symbol,
                "trades": 0,
                "total_pnl": 0.0,
                "avg_return_pct": 0.0,
            },
        )
        symbol_payload["trades"] += 1
        symbol_payload["total_pnl"] += float(item["pnl"])
        symbol_payload["avg_return_pct"] += float(item["return_pct"])

    normalized = []
    for payload in by_symbol.values():
        trades = max(int(payload["trades"]), 1)
        normalized.append(
            {
                "symbol": payload["symbol"],
                "stock_name": payload["stock_name"],
                "trades": trades,
                "total_pnl": round(payload["total_pnl"], 2),
                "avg_return_pct": round(payload["avg_return_pct"] / trades, 4),
            }
        )

    winners = sorted(normalized, key=lambda item: item["total_pnl"], reverse=True)[:3]
    losers = sorted(normalized, key=lambda item: item["total_pnl"])[:3]
    return {"winners": winners, "losers": losers}


def _build_trade_ai_context(
    diagnostics: TradeDiagnosticsResponse,
    trades: list[dict[str, Any]],
    round_trips: list[dict[str, Any]],
) -> dict[str, Any]:
    active_days = sorted(
        {
            trade["trade_date"] if isinstance(trade["trade_date"], str) else trade["trade_date"].isoformat()
            for trade in trades
        }
    )
    trade_frequency = round(len(trades) / max(len(active_days), 1), 2) if trades else 0.0
    symbol_summary = _round_trip_symbol_summary(round_trips)
    return {
        "coverage_text": diagnostics.coverage_text,
        "latest_batch": (
            {
                "batch_id": diagnostics.latest_batch.batch_id,
                "broker": diagnostics.latest_batch.broker,
                "imported_count": diagnostics.latest_batch.imported_count,
                "symbol_count": diagnostics.latest_batch.symbol_count,
                "start_date": diagnostics.latest_batch.start_date.isoformat()
                if diagnostics.latest_batch.start_date
                else None,
                "end_date": diagnostics.latest_batch.end_date.isoformat()
                if diagnostics.latest_batch.end_date
                else None,
            }
            if diagnostics.latest_batch
            else None
        ),
        "style_profile": {
            "display_name": diagnostics.style_profile.display_name,
            "confidence": diagnostics.style_profile.confidence,
            "summary": diagnostics.style_profile.summary,
            "traits": diagnostics.style_profile.traits,
        },
        "summary_metrics": [
            {"label": item.label, "value": item.value, "detail": item.detail}
            for item in diagnostics.summary_metrics
        ],
        "win_loss_comparison": [
            {"label": item.label, "value": item.value, "detail": item.detail}
            for item in diagnostics.win_loss_comparison
        ],
        "error_patterns": [
            {"title": item.title, "detail": item.detail, "severity": item.severity}
            for item in diagnostics.error_patterns
        ],
        "effective_patterns": [
            {"title": item.title, "detail": item.detail, "severity": item.severity}
            for item in diagnostics.effective_patterns
        ],
        "recommendations": diagnostics.recommendations,
        "trade_meta": {
            "trade_count": len(trades),
            "closed_round_trip_count": len(round_trips),
            "trade_frequency_per_active_day": trade_frequency,
            "top_winners": symbol_summary["winners"],
            "top_losers": symbol_summary["losers"],
        },
    }


def _fallback_trade_ai_analysis(diagnostics: TradeDiagnosticsResponse) -> dict[str, Any]:
    style = diagnostics.style_profile.display_name if diagnostics.style_profile else "待判断"
    strengths = [item.title for item in diagnostics.effective_patterns[:3]] or ["当前先按结构化统计结果观察有效习惯。"]
    weaknesses = [item.title for item in diagnostics.error_patterns[:3]] or ["当前未识别到足够明确的高频错误模式。"]
    adjustments = diagnostics.recommendations[:4] or ["先继续积累更完整的交割单，再做更强的 AI 复盘。"]
    next_cycle_plan = adjustments[:3]
    behavior_tags = diagnostics.style_profile.traits[:3] if diagnostics.style_profile else ["样本不足"]
    return {
        "status": "rule_stub_ready",
        "model": None,
        "confidence": 0.66,
        "summary": "先按结构化交易诊断结论执行，暂不额外放大主观判断。",
        "trader_profile": f"当前更接近 {style}。",
        "strengths": strengths,
        "weaknesses": weaknesses,
        "behavior_tags": behavior_tags,
        "adjustments": adjustments,
        "next_cycle_plan": next_cycle_plan,
        "source": "rules",
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


def _fuse_trade_ai_analysis(
    diagnostics: TradeDiagnosticsResponse,
    ai_result: dict[str, Any],
    fallback_analysis: dict[str, Any],
) -> dict[str, Any]:
    confidence = float(ai_result.get("confidence", fallback_analysis.get("confidence", 0.7)))
    confidence = max(0.0, min(1.0, confidence))
    return {
        "status": "ai_live",
        "model": ai_result.get("model") or _settings().gemini_model,
        "confidence": round(confidence, 2),
        "summary": ai_result.get("summary") or fallback_analysis["summary"],
        "trader_profile": ai_result.get("trader_profile") or fallback_analysis["trader_profile"],
        "strengths": list(ai_result.get("strengths") or fallback_analysis["strengths"])[:4],
        "weaknesses": list(ai_result.get("weaknesses") or fallback_analysis["weaknesses"])[:4],
        "behavior_tags": list(ai_result.get("behavior_tags") or fallback_analysis["behavior_tags"])[:4],
        "adjustments": list(ai_result.get("adjustments") or diagnostics.recommendations or fallback_analysis["adjustments"])[:5],
        "next_cycle_plan": list(ai_result.get("next_cycle_plan") or fallback_analysis["next_cycle_plan"])[:3],
        "source": "gemini",
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


def get_effective_trade_diagnostics_ai(
    batch_id: Optional[str],
    diagnostics: TradeDiagnosticsResponse,
) -> Optional[dict[str, Any]]:
    if not batch_id:
        return _fallback_trade_ai_analysis(diagnostics)
    cached = _get_cached_trade_ai_analysis(batch_id)
    if cached is not None:
        return cached
    return _fallback_trade_ai_analysis(diagnostics)


def regenerate_trade_diagnostics_ai_analysis(batch_id: Optional[str] = None) -> dict[str, Any]:
    batch = _load_batch_by_id(batch_id) if batch_id else _load_latest_batch()
    if batch is None:
        diagnostics = _build_demo_diagnostics()
        return _fallback_trade_ai_analysis(diagnostics)

    trades = _load_trades_for_batch(batch.batch_id)
    round_trips = _compute_round_trips(trades)
    summary_metrics, metrics = _build_metric_cards(round_trips, trades)
    style_profile = _infer_style(round_trips, trades)
    win_loss_comparison = _build_win_loss_comparison(round_trips, metrics)
    error_patterns, effective_patterns, recommendations = _build_insights(metrics, trades)
    diagnostics = TradeDiagnosticsResponse(
        status="live",
        account_label=f"{batch.broker} 最近导入批次",
        coverage_text=f"覆盖 {batch.start_date} 至 {batch.end_date}，共 {batch.imported_count} 条成交记录，{batch.symbol_count} 只股票。",
        latest_batch=batch,
        summary_metrics=summary_metrics,
        style_profile=style_profile,
        win_loss_comparison=win_loss_comparison,
        error_patterns=error_patterns,
        effective_patterns=effective_patterns,
        recommendations=recommendations,
        recent_batches=_load_recent_batches(limit=5),
    )
    fallback = _fallback_trade_ai_analysis(diagnostics)
    context = _build_trade_ai_context(diagnostics, trades, round_trips)
    try:
        ai_result = call_gemini_trade_diagnostics_analysis(context)
        fused = _fuse_trade_ai_analysis(diagnostics, ai_result, fallback)
    except GeminiClientError as exc:
        fused = fallback
        fused["status"] = "ai_error"
        fused["adjustments"] = [*fused.get("adjustments", [])[:3], str(exc)]
    _save_cached_trade_ai_analysis(batch.batch_id, fused)
    return fused


def _build_live_trade_diagnostics(batch: ImportBatchSummary) -> TradeDiagnosticsResponse:
    trades = _load_trades_for_batch(batch.batch_id)
    round_trips = _compute_round_trips(trades)
    summary_metrics, metrics = _build_metric_cards(round_trips, trades)
    style_profile = _infer_style(round_trips, trades)
    win_loss_comparison = _build_win_loss_comparison(round_trips, metrics)
    error_patterns, effective_patterns, recommendations = _build_insights(metrics, trades)
    diagnostics = TradeDiagnosticsResponse(
        status="live",
        account_label=f"{batch.broker} 最近导入批次",
        coverage_text=f"覆盖 {batch.start_date} 至 {batch.end_date}，共 {batch.imported_count} 条成交记录，{batch.symbol_count} 只股票。",
        latest_batch=batch,
        summary_metrics=summary_metrics,
        style_profile=style_profile,
        win_loss_comparison=win_loss_comparison,
        error_patterns=error_patterns,
        effective_patterns=effective_patterns,
        recommendations=recommendations,
        recent_batches=_load_recent_batches(limit=5),
    )
    diagnostics.ai_analysis = (
        TradeDiagnosticsAIResponse(**get_effective_trade_diagnostics_ai(batch.batch_id, diagnostics))
        if batch.batch_id
        else None
    )
    return diagnostics


def get_trade_diagnostics() -> TradeDiagnosticsResponse:
    ensure_trade_tables()
    latest_batch = _load_latest_batch()
    if latest_batch is None:
        diagnostics = _build_demo_diagnostics()
        diagnostics.ai_analysis = TradeDiagnosticsAIResponse(
            **_fallback_trade_ai_analysis(diagnostics)
        )
        return diagnostics

    return _build_live_trade_diagnostics(latest_batch)


def export_trade_schema_blueprint() -> dict[str, Any]:
    ensure_trade_tables()
    return {
        "tables": {
            "import_batches": [
                "batch_id TEXT PRIMARY KEY",
                "imported_at TEXT NOT NULL",
                "broker TEXT NOT NULL",
                "source_type TEXT NOT NULL",
                "filename TEXT NOT NULL",
                "detected_format TEXT NOT NULL",
                "row_count INTEGER NOT NULL",
                "imported_count INTEGER NOT NULL",
                "ignored_count INTEGER NOT NULL",
                "symbol_count INTEGER NOT NULL",
                "start_date TEXT",
                "end_date TEXT",
                "notes TEXT",
            ],
            "trades": [
                "id INTEGER PRIMARY KEY AUTOINCREMENT",
                "batch_id TEXT NOT NULL",
                "trade_date TEXT NOT NULL",
                "trade_time TEXT",
                "symbol TEXT NOT NULL",
                "stock_name TEXT",
                "market TEXT",
                "side TEXT NOT NULL",
                "quantity INTEGER NOT NULL",
                "price REAL NOT NULL",
                "amount REAL NOT NULL",
                "commission REAL NOT NULL DEFAULT 0",
                "stamp_tax REAL NOT NULL DEFAULT 0",
                "transfer_fee REAL NOT NULL DEFAULT 0",
                "other_fee REAL NOT NULL DEFAULT 0",
                "net_amount REAL NOT NULL",
                "account_masked TEXT",
                "broker TEXT NOT NULL",
                "source_type TEXT NOT NULL",
                "raw_row_id INTEGER",
            ],
        },
        "profiles": json.loads(get_import_profiles().model_dump_json()),
    }
